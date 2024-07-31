import openpyxl
import win32com.client as win32
import pandas as pd
import os
import sys

def get_desktop_path(filename):
    if filename in ["calcurate.xlsm", "default.hwp"]:
        return os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "StructFlow-Automator-Private", "References", filename
        )
    elif filename in ["temp_excel.xlsx", "temp_hwp.hwp", "temp_all_data_excel.xlsx"]:
        if os.path.exists(os.path.join(os.path.expanduser("~"), "OneDrive")):
            if os.path.exists(os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")):
                return os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", filename)
            else:
                return os.path.join(os.path.expanduser("~"), "OneDrive", "바탕 화면", filename)
        else:
            if os.path.exists(os.path.join(os.path.expanduser("~"), "Desktop")):
                return os.path.join(os.path.expanduser("~"), "Desktop", filename)
            else:
                return os.path.join(os.path.expanduser("~"), "바탕 화면", filename)

    return os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "temp_dir", filename
    )

def ensure_directory_exists(path):
    directory = os.path.dirname(path)
    if not os.path.exists(directory):
        os.makedirs(directory)
        print(f"Created directory: {directory}")

def check_file_exists(path):
    if not os.path.exists(path):
        print(f"Error: File not found: {path}")
        return False
    return True

class ExcelProcessor:
    def __init__(self, excel_path, temp_excel_path):
        self.excel_path = excel_path
        self.temp_excel_path = temp_excel_path

    def process_data(self, data1_path, data2_path, data3_path):
        # 파일 존재 여부 확인
        if not all(map(check_file_exists, [data1_path, data2_path, data3_path, self.excel_path])):
            return

        modified_data2 = self.modify_data2(data2_path)

        try:
            wb = openpyxl.load_workbook(self.excel_path, keep_vba=True)
        except Exception as e:
            print(f"Failed to load workbook: {e}", flush=True)
            return

        code_sheet = wb["code"]
        self.process_data1_and_data2(data1_path, modified_data2, code_sheet)

        section_sheet = wb["section"]
        self.process_data3(data3_path, section_sheet)

        updated_excel_path = self.excel_path.replace(".xlsm", "_updated.xlsm")

        try:
            wb.save(updated_excel_path)
        except Exception as e:
            print(f"Failed to save updated Excel file: {e}", flush=True)
        finally:
            wb.close()

        try:
            excel = win32.gencache.EnsureDispatch("Excel.Application")
            wb = excel.Workbooks.Open(updated_excel_path)
            excel.Calculate()
            wb.Save()
            wb.Close()
            excel.Quit()
        except Exception as e:
            print(f"Failed to open and save Excel file with Excel COM: {e}", flush=True)
            if "excel" in locals():
                excel.Quit()
            return

        try:
            wb = openpyxl.load_workbook(updated_excel_path, data_only=True)
        except Exception as e:
            print(f"Failed to load workbook for data extraction: {e}", flush=True)
            return

        summary_sheet = wb["종합"]

        table_data = []
        for row in summary_sheet.iter_rows(values_only=True):
            if row[0] is not None:
                filtered_row = [
                    row[0],
                    row[1],
                    row[2],
                    row[7],
                ]
                filtered_row = [
                    value if value is not None else "" for value in filtered_row
                ]
                table_data.append(filtered_row)

        df = pd.DataFrame(table_data)

        ensure_directory_exists(self.temp_excel_path)
        try:
            df.to_excel(self.temp_excel_path, index=False, header=False)
            print(f"Temporary Excel file created successfully: {self.temp_excel_path}")
        except Exception as e:
            print(f"Failed to save temporary Excel file: {e}", flush=True)
            return

    def modify_data2(self, data2_path):
        with open(data2_path, "r", encoding="utf-8") as file:
            lines = file.readlines()
        if not lines:
            print(f"Warning: {data2_path} is empty")
            return []
        header = lines[0].strip().split("\t")
        if "Pu" not in header:
            print(f"Warning: Column 'Pu' not found in header of {data2_path}")
            return lines
        pu_index = header.index("Pu")
        header.insert(pu_index + 1, header[pu_index])
        modified_data = ["\t".join(header)]
        for line in lines[1:]:
            values = line.strip().split("\t")
            if len(values) > pu_index:
                values.insert(pu_index + 1, values[pu_index])
            modified_data.append("\t".join(values))
        return modified_data

    def clean_decimal(self, value):
        try:
            value = str(value)
            float_value = float(value)
            if float_value.is_integer():
                return int(float_value)
            return float(f"{float_value:.4f}".rstrip("0").rstrip("."))
        except ValueError:
            return value

    def process_data1_and_data2(self, data1_path, modified_data2, code_sheet):
        self.process_data_file(data1_path, code_sheet, start_row=3)
        self.process_data_list(modified_data2, code_sheet, start_row=self.get_last_row(code_sheet) + 1)

    def process_data3(self, data3_path, section_sheet):
        self.process_data_file(data3_path, section_sheet, start_row=1)

    def process_data_file(self, file_path, sheet, start_row):
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                data = file.readlines()
        except FileNotFoundError:
            print(f"File not found: {file_path}", flush=True)
            return
        except Exception as e:
            print(f"Failed to read file {file_path}: {e}", flush=True)
            return

        for row_idx, line in enumerate(data, start=start_row):
            values = line.strip().split("\t")
            for col_idx, value in enumerate(values, start=1):
                cleaned_value = self.clean_decimal(value)
                sheet.cell(row=row_idx, column=col_idx, value=cleaned_value)

    def process_data_list(self, data_list, sheet, start_row):
        for row_idx, line in enumerate(data_list, start=start_row):
            values = line.strip().split("\t")
            for col_idx, value in enumerate(values, start=1):
                cleaned_value = self.clean_decimal(value)
                sheet.cell(row=row_idx, column=col_idx, value=cleaned_value)

    def get_last_row(self, sheet):
        last_row = sheet.max_row
        while last_row > 0 and all(
            sheet.cell(row=last_row, column=col).value is None
            for col in range(1, sheet.max_column + 1)
        ):
            last_row -= 1
        return last_row

    def set_clipboard_from_excel(self):
        if not check_file_exists(self.temp_excel_path):
            return

        try:
            excel = win32.Dispatch("Excel.Application")
            workbook = excel.Workbooks.Open(self.temp_excel_path)
            worksheet = workbook.Worksheets(1)
            worksheet.UsedRange.Copy()
            workbook.Close(False)
            excel.Quit()
            print("Data copied to clipboard from temporary Excel file.", flush=True)
        except Exception as e:
            print(f"Error copying data to clipboard: {e}", flush=True)

def main():
    data1_path = get_desktop_path("data_1.txt")
    data2_path = get_desktop_path("data_2.txt")
    data3_path = get_desktop_path("data_3.txt")
    excel_path = get_desktop_path("calcurate.xlsm")
    temp_excel_path = get_desktop_path("temp_excel.xlsx")

    # 파일 경로 출력
    print(f"data1_path: {data1_path}")
    print(f"data2_path: {data2_path}")
    print(f"data3_path: {data3_path}")
    print(f"excel_path: {excel_path}")
    print(f"temp_excel_path: {temp_excel_path}")

    excel_processor = ExcelProcessor(excel_path, temp_excel_path)
    excel_processor.process_data(data1_path, data2_path, data3_path)

    for _ in range(3):  # 3번 반복
        excel_processor.set_clipboard_from_excel()
        input("데이터를 붙여넣기 한 후 Enter 키를 눌러주세요...")

if __name__ == "__main__":
    main()